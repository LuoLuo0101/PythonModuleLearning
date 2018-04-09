# -*- coding: utf-8 -*-
# @Time    : 18-4-4 下午2:13
# @Author  : luo
# @File    : 05.测试用例.py
# @Software: PyCharm

import openpyxl
from datetime import datetime

from openpyxl.styles import Alignment, Font

"""
B2-Q13：第一个表格区域

B2-L2：左边标题（交易审批—日数据）
M2-Q2：右边标题（交易审批—累计数据）
业务类型(B3)	渠道(C3)	企业名称(D3)	申请件数(E3)	通过件数(F3)	通过率(G3)	申请金额(H3)	通过金额(I3)	免面签率(J3)
首次进件企业数	(K3) 企业是否首次进件(L3)	累计通过金额(M3)	期末贷款余额(N3)	企业授信余额(O3)	逾期率(P)	不良率(Q3)
"""


class ExcelCreater(object):

    def __init__(self, file_name, suffix="xlsx"):
        self.file_name = file_name
        self.suffix = suffix
        self.wb = openpyxl.Workbook(write_only=False, iso_dates=True)
        self.ws = self.wb.active

    def change_file_name(self, file_name=None):
        self.file_name = file_name

    def change_suffix(self, suffix="xlsx"):
        self.suffix = suffix

    def set_title(self, title=None):
        self.ws.title = title

    def create_sheet(self, sheet_name=None, index=None):
        self.ws = self.wb.create_sheet(title=sheet_name, index=index)

    def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
        self.ws.merge_cells(
            range_string=range_string,
            start_row=start_row,
            start_column=start_column,
            end_row=end_row,
            end_column=end_column
        )

    def set_value(self, cell_name=None, row=None, column=None, value=None):
        align = Alignment(horizontal='center', vertical='center')
        font = Font(size=10)
        if cell_name is not None:
            self.ws[cell_name] = value
            self.ws[cell_name].alignment = align
            self.ws[cell_name].font = font
        elif isinstance(row, int) and isinstance(column, int):
            self.ws.cell(row=row, column=column, value=value)
            self.ws.cell(row=row, column=column).alignment = align
            self.ws.cell(row=row, column=column).font = font
        else:
            return False
        return True

    def set_center_alignment(self, cell_name=None, row=None, column=None):
        align = Alignment(horizontal='center', vertical='center')
        if cell_name is not None:
            self.ws[cell_name].alignment = align
        elif isinstance(row, int) and isinstance(column, int):
            self.ws.cell(row=row, column=column).alignment = align
        else:
            return False
        return True

    def adjust_column(self):
        dims = {}
        for row in self.ws.rows:
            for cell in row:
                if cell.value:
                    calc_len = 0
                    z_count = 0
                    for c in str(cell.value):
                        if c.isalpha():
                            z_count += 1
                        else:
                            calc_len += 1
                    calc_len += int(z_count * 35 / 19)
                    dims[cell.column] = max((dims.get(cell.column, 0), calc_len + 4))
        for col, value in dims.items():
            self.ws.column_dimensions[col].width = value

    def save(self):
        file_name = "%s.%s" % (self.file_name, self.suffix)
        self.wb.save(filename=file_name)


if __name__ == '__main__':
    c_chr = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U",
             "V", "W", "X", "Y", "Z"]
    date_now = datetime.now().date()
    title = date_now.strftime("%Y-%m-%d")
    excel = ExcelCreater(file_name=title)  # 设置
    excel.set_title(title=title)  # 设置sheet标题

    excel.merge_cells(range_string="B2:L2")
    excel.merge_cells(range_string="M2:Q2")
    # 两个标题
    title_dict = {
        "B2": "交易审批—日数据",
        "M2": "交易审批—累计数据"
    }
    for k, v in title_dict.items():
        excel.set_value(cell_name=k, value=v)

    # 每一个列标题
    set_value_dict = {
        "B3": "业务类型",
        "C3": "渠道",
        "D3": "企业名称",
        "E3": "申请件数",
        "F3": "通过件数",
        "G3": "通过率",
        "H3": "申请金额",
        "I3": "通过金额",
        "J3": "免面签率",
        "K3": "首次进件企业数",
        "L3": "企业是否首次进件",
        "M3": "累计通过金额",
        "N3": "期末贷款余额",
        "O3": "企业授信余额",
        "P3": "逾期率",
        "Q3": "不良率",
    }
    for k, v in set_value_dict.items():
        excel.set_value(cell_name=k, value=v)

    enterprise_datas = [  # TODO 需要动态更换的数据源
        {
            "enterprise_name": "北京中诺口腔医院",
            "apply_piece": 1,
            "pass_piece": 0,
            "pass_rate": "=%s%d/%s%d",
            "apply_sum": 12697.30,
            "pass_sum": 0.00,
            "no_face_rate": "0.00%",
            "first_in_count": 0,
            "is_first_in": 0
        },
        {
            "enterprise_name": "北京倪氏威尔默口腔",
            "apply_piece": 1,
            "pass_piece": 1,
            "pass_rate": "=%s%d/%s%d",
            "apply_sum": 7220.00,
            "pass_sum": 7220.00,
            "no_face_rate": "0.00%",
            "first_in_count": 0,
            "is_first_in": 0
        },
        {
            "enterprise_name": "北京昌通军程医药科技发展天通苑口腔诊所",
            "apply_piece": 1,
            "pass_piece": 0,
            "pass_rate": "=%s%d/%s%d",
            "apply_sum": 19000.00,
            "pass_sum": 0.00,
            "no_face_rate": "0.00%",
            "first_in_count": 0,
            "is_first_in": 1
        },
        {
            "enterprise_name": "北京维尔口腔医院上地口腔",
            "apply_piece": 1,
            "pass_piece": 1,
            "pass_rate": "=%s%d/%s%d",
            "apply_sum": 18400.00,
            "pass_sum": 18400.00,
            "no_face_rate": "100.00%",
            "first_in_count": 0,
            "is_first_in": 0
        },
        {
            "enterprise_name": "太原市恒伦口腔医院",
            "apply_piece": 1,
            "pass_piece": 1,
            "pass_rate": "=%s%d/%s%d",
            "apply_sum": 32200.00,
            "pass_sum": 32200.00,
            "no_face_rate": "0.00%",
            "first_in_count": 0,
            "is_first_in": 0
        }
    ]
    r, c = 4, 4
    for item in enterprise_datas:
        enterprise_name = item.get("enterprise_name")
        excel.set_value(cell_name="%s%d" % (c_chr[c - 1], r), value=enterprise_name)
        c += 1

        apply_piece = item.get("apply_piece")
        excel.set_value(cell_name="%s%d" % (c_chr[c - 1], r), value=apply_piece)
        c += 1

        pass_piece = item.get("pass_piece")
        excel.set_value(cell_name="%s%d" % (c_chr[c - 1], r), value=pass_piece)
        c += 1

        pass_rate = item.get("pass_rate") % (c_chr[c - 2], r, c_chr[c - 3], r)
        excel.set_value(cell_name="%s%d" % (c_chr[c - 1], r), value=pass_rate)
        c += 1

        apply_sum = item.get("apply_sum")
        excel.set_value(cell_name="%s%d" % (c_chr[c - 1], r), value=apply_sum)
        c += 1

        pass_sum = item.get("pass_sum")
        excel.set_value(cell_name="%s%d" % (c_chr[c - 1], r), value=pass_sum)
        c += 1

        no_face_rate = item.get("no_face_rate")
        excel.set_value(cell_name="%s%d" % (c_chr[c - 1], r), value=no_face_rate)
        c += 1

        first_in_count = item.get("first_in_count")
        excel.set_value(cell_name="%s%d" % (c_chr[c - 1], r), value=first_in_count if first_in_count else "")
        c += 1

        is_first_in = item.get("is_first_in")
        excel.set_value(cell_name="%s%d" % (c_chr[c - 1], r), value="是" if is_first_in else "否")
        c += 1

        r += 1
        c = 4

    excel.merge_cells(range_string="B4:B%d" % (4 + len(enterprise_datas) - 1))
    excel.set_value(cell_name="B4", value="大众版")

    excel.merge_cells(range_string="C4:C%d" % (4 + len(enterprise_datas) - 1))
    excel.set_value(cell_name="C4", value="金服侠")

    c = 2  # B
    # 大众版小计
    excel.merge_cells(range_string="B%d:D%d" % (r, r))
    excel.set_value(cell_name="B%d" % r, value="大众版小计")

    # 大众版申请件数总和求和
    excel.set_value(cell_name="E%d" % r, value="=SUM(E4:E%d)" % (r - 1))
    # 大众版通过件数总和求和
    excel.set_value(cell_name="F%d" % r, value="=SUM(F4:F%d)" % (r - 1))
    # 大众版平均通过率
    excel.set_value(cell_name="G%d" % r, value="=F%d/E%d" % (r, r))
    # 大众版申请金额求和
    excel.set_value(cell_name="H%d" % r, value="=SUM(H4:H%d)" % (r - 1))
    # 大众版通过金额求和
    excel.set_value(cell_name="I%d" % r, value="==SUM(I4:I%d)" % (r - 1))
    # 大众版平均免面签率
    excel.set_value(cell_name="J%d" % r, value="=1/F%d" % r)
    # 大众版首次进件企业数求和
    excel.set_value(cell_name="K%d" % r, value="=SUM(K4:K%d)" % (r - 1))
    # 大众版企业是否首次进件
    excel.set_value(cell_name="L%d" % r, value="")
    excel.set_value(cell_name="D11", value="00000000000000000000000000000000")

    # 调整宽度
    excel.adjust_column()

    excel.save()
