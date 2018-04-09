# -*- coding: utf-8 -*-
# @Time    : 18-4-3 下午3:20
# @Author  : luo
# @File    : 02.单元格操作.py
# @Software: PyCharm

import openpyxl
from openpyxl.drawing.image import Image

# 工作薄
wb = openpyxl.load_workbook(
    filename="1.xlsx",
    read_only=False
)

# 获取一个 sheet
ws = wb.get_sheet_by_name(name="工作页04")

# 获取单元格的值的三种方式（Cell对象）
# 第一种方式
a_value1 = ws["A2"].value
print(a_value1)

# 第二种方式
# or 非常推荐 遍历都很方便
a_value2 = ws.cell(row=1, column=1).value
print(a_value2)

# 第三种方式：获取到范围内的Cell对象的tuple阵列((Cell, Cell, Cell), (Cell, Cell, Cell), (Cell, Cell, Cell))
cells = ws['A1': 'E4']
print(cells)

# 获取C那一列
colC = ws['C']
# 获取CD两列，前后都包括
col_range = ws['C:D']
# 获取第十行
row10 = ws[10]
# 获取5-10行，前后都包括
row_range = ws[5:10]

# 为单元格赋值两种方式
# 第一种方式
ws['B2'] = "alic"

# 第二种方式
ws.cell(row=1, column=2, value="123")

# 遍历多个单元格
# 第一种遍历方式：这将会创建100*100个空的cell
for i in range(1, 101):
    for j in range(1, 101):
        ws.cell(row=i, column=j)

# 第二种遍历方式
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)

# 第三种遍历方式
for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
    for cell in col:
        print(cell)

# 遍历所有行
ws_rows = tuple(ws.rows)

# 遍历所有列
ws_cols = tuple(ws.columns)

# 方式一：合并/取消合并单元格
ws.merge_cells("A1:B2")
ws.unmerge_cells("A1:B2")

# 方式二：合并/取消合并单元格
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
ws.unmerge_cells(start_row=2, start_column=1, end_row=2, end_column=4)

# 创建一个图片对象
img = Image('1.jpg')
# 将图片插入某个单元格位置，并不会替换原来单元格的数据
ws.add_image(img, 'A1')


# 折叠列，将C和D列的宽度变成0，变成隐藏状态
ws.column_dimensions.group('C', 'D', hidden=True)

# 添加一/多行数据
ws6 = wb.get_sheet_by_name(name="工作页06")

row = ["123", "456", "789"]

ws6.append(row)

# 无效果
# rows = [("10123", "10456", "10789"),
#         ("10123", "10456", "10789")]
#
# list(zip(*rows))
#
# ws6.append(row)

wb.save(filename="1.xlsx")
