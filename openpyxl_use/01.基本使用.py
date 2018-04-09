# -*- coding: utf-8 -*-
# @Time    : 18-4-3 下午2:01
# @Author  : luo
# @File    : 01.基本使用.py
# @Software: PyCharm

import openpyxl

"""
https://www.jianshu.com/p/50b075c5f5e7

指定了 write_only=True，操作对象是，只写对象
    openpyxl.worksheet.write_only.WriteOnlyWorksheet

指定了 read_only=True，操作对象是，只读对象
    openpyxl.worksheet.read_only.ReadOnlyWorksheet
    
两个都没用指定，操作对象是，可读写对象
    <openpyxl.workbook.workbook.Workbook object at 0x7f72399822b0>

openpyxl不能读取Excle中所有的对象，当打开和保存相同名字的文件的时候，图片和图表将会丢失

公式必须使用英文名，并且公式的参数必须使用逗号分隔，不能使用其他的符号如分号
"""

# 创建一个工作薄
wb0 = openpyxl.Workbook(
    write_only=True,    # 是否以只读形式打开 openpyxl.worksheet.write_only.WriteOnlyWorksheet 对象
    iso_dates=True
)

filepath = "1.xlsx"

# 打开一个已经存在的工作薄
wb1 = openpyxl.load_workbook(
    filename=filepath,      # 打开的文件名
    read_only=False,        # 只读模式打开，返回的是 openpyxl.worksheet.read_only.ReadOnlyWorksheet 对象
    guess_types=True,       # 格式以及类型推断，'31.50' 保存成 31.5推断数据类型、'12%' 保存成 0.12、时间格式也是
)

# 选中 sheet 的四种方式
# 第一种方式：激活默认工作sheet
ws1 = wb1.active

# 第二种方式：通过索引加载sheet index从0开始
wss = wb1.worksheets     # 获取当前文件所有 sheet 的 list
ws2 = wss[0]

# 第三种方式：通过 sheet 名选中 sheet
ws3 = wb1.get_sheet_by_name(name="工作页0001")
print(ws3.title)

# 第四种方式：直接通过对 WookBook 进行操作
ws4 = wb1["工作页0001"]


# 获取 sheet 名字的两种方式
# 第一种方式：获取当前 sheet 的名字
ws_title = ws4.title
print(ws_title)

# 第二种方式：遍历 sheet 名字列表 list
sheet_names = wb1.get_sheet_names()
print(sheet_names)

# 第三种方式：遍历 sheet 名字列表 list
sheet_names2 = wb1.sheetnames
print(sheet_names2)

# 修改 sheet 名字
ws3.title = "工作页0001"

ws_rows = ws3.rows
ws_columns = ws3.columns

print(ws_rows)
print(ws_columns)

row_len = len(list(ws_rows))        # 数据行数
column_len = len(list(ws_columns))  # 数据列数

print(row_len)
print(column_len)

# 获取 sheet 的最大行数
print(ws3.max_row)
print(ws3.max_column)

# 创建 sheet
# 默认插入在工作薄的末尾
ws = wb1.create_sheet(
    title="工作页06",    # 工作页标题
    index=1             # 索引从0开始算，工作页插入的位置，如果该数字超过存在的最大索引，那么默认插入到最后
)

# 复制一个sheet
target = wb1.copy_worksheet(ws)
target.title = "我是复制的sheet"

# 更改sheet tab 的颜色 RRGGBB
target.sheet_properties.tabColor = "FF0000"

# 删除一个sheet，如果该sheet不存在会报错
wb1.remove(wb1.get_sheet_by_name('one'))

# 保存工作薄
# 直接保存为该文件名
# 如果添加 wb1.template = True 保存为模板
wb1.template = True
wb1.save(
    filename=filepath   # 保存路径+文件名
)
