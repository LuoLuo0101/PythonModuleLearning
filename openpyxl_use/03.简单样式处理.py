# -*- coding: utf-8 -*-
# @Time    : 18-4-3 下午5:17
# @Author  : luo
# @File    : 03.简单样式处理.py
# @Software: PyCharm

import openpyxl

from openpyxl.styles import Alignment, Font

# 工作薄
wb = openpyxl.load_workbook(
    filename="日报_2018-03-29.xlsx",
    read_only=False
)

# 获取一个 sheet
ws = wb.active

"""
# 文本对齐方式
# 居中对齐
align = Alignment(horizontal='center', vertical='center')
ws.cell(row=1, column=1).alignment = align

# 字体大小
font = Font(size=10)
ws.cell(row=1, column=1).font = font
"""
cell1 = ws["A1"]
col = ws.column_dimensions['B']
row = ws.row_dimensions['2']
print(col)
print(row)
# bgColor：rgb/value 00000000
# fgColor：rgb/value 00000000
# fond：name 宋体 sz 11
wb.save(filename="1.xlsx")
